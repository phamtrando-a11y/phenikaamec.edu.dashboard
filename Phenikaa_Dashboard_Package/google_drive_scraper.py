import os
import re
import io
import json
import requests
import pandas as pd
import numpy as np
import unicodedata
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

def normalize_text_std(text):
    if not isinstance(text, str):
        return ""
    return unicodedata.normalize('NFC', text).strip().lower()

def clean_program_name(name):
    if not isinstance(name, str):
        return ""
    name = normalize_text_std(name)
    name = re.sub(r'\(.*?\)', '', name)
    name = re.sub(r'\[.*?\]', '', name)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

def get_urls_from_cell(val, drive_service=None):
    if pd.isna(val):
        return []
    val_str = str(val).strip()
    urls = re.findall(r'https?://[^\s,;"\'\)]+', val_str)
    if not urls and drive_service and len(val_str) > 15:
        if val_str.lower() not in ['không có', 'chưa có', 'không', 'chưa', 'nan', 'none']:
            try:
                val_base = val_str
                if val_base.lower().endswith('.xlsx'):
                    val_base = val_base[:-5]
                elif val_base.lower().endswith('.csv'):
                    val_base = val_base[:-4]
                
                query = f"(name='{val_base}' or name='{val_base}.xlsx' or name='{val_base}.csv' or name='{val_str}') and trashed=false"
                results = drive_service.files().list(q=query, fields="files(id, name)").execute()
                items = results.get('files', [])
                if items:
                    urls = [f"https://docs.google.com/spreadsheets/d/{items[0]['id']}/edit"]
            except:
                pass
    return urls

def get_download_url(url):
    sheet_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if sheet_match:
        sp_id = sheet_match.group(1)
        res_key = ""
        rk_match = re.search(r'(resourcekey=[a-zA-Z0-9-_]+)', url)
        if rk_match:
            res_key = "&" + rk_match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sp_id}/export?format=xlsx{res_key}", "google_sheet"
        
    drive_match = re.search(r'/file/d/([a-zA-Z0-9-_]+)', url)
    if drive_match:
        file_id = drive_match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}", "google_drive"
        
    if '1drv.ms' in url or 'sharepoint.com' in url:
        dl_url = url + ('&download=1' if '?' in url else '?download=1')
        return dl_url, "onedrive"
        
    return None, None

def find_header_and_data(df):
    best_row_idx = 0
    max_matches = 0
    keywords = {'dấu thời gian', 'timestamp', 'điểm số', 'score', 'họ tên', 'họ và tên', 'tên', 
                'mã nhân viên', 'mã nhân sự', 'mã nv', 'mã ns', 'email', 'địa chỉ email'}
    
    for idx, row in df.iterrows():
        row_str = [str(val).lower() for val in row.values]
        matches = sum(1 for kw in keywords if any(kw in rv for rv in row_str))
        if matches > max_matches:
            max_matches = matches
            best_row_idx = idx
            
    if max_matches >= 2:
        columns = df.iloc[best_row_idx].values
        cleaned_columns = []
        for c in columns:
            c_str = str(c).strip()
            if not c_str or c_str.lower() == 'nan':
                cleaned_columns.append(f"Unnamed_{len(cleaned_columns)}")
            else:
                cleaned_columns.append(c_str)
        df_clean = df.iloc[best_row_idx+1:].copy()
        df_clean.columns = cleaned_columns
        return df_clean
    return df

def map_columns(df):
    cols = df.columns
    mapped = {'name_col': None, 'id_col': None, 'score_col': None, 'email_col': None}
    
    for col in cols:
        col_norm = normalize_text_std(col)
        if not mapped['name_col']:
            if col_norm in ('họ tên', 'họ và tên', 'họ tên đầy đủ', 'họ và tên đầy đủ', 'tên', 'name', 'nhân sự', 'nguyễn văn a', 'họ tên?'):
                mapped['name_col'] = col
            elif 'họ' in col_norm and 'tên' in col_norm:
                mapped['name_col'] = col
        if not mapped['id_col']:
            if col_norm in ('mã nhân sự', 'mã nhân viên', 'mã nv', 'mã ns', 'msnv', 'msns', 'id', 'mã số nhân viên'):
                mapped['id_col'] = col
            elif 'mã' in col_norm and ('nhân viên' in col_norm or 'nhân sự' in col_norm or 'nv' in col_norm or 'ns' in col_norm):
                mapped['id_col'] = col
        if not mapped['score_col']:
            if col_norm in ('điểm số', 'điểm', 'score', 'kết quả'):
                mapped['score_col'] = col
            elif 'điểm' in col_norm or 'score' in col_norm:
                mapped['score_col'] = col
        if not mapped['email_col']:
            if col_norm in ('email', 'địa chỉ email', 'email address'):
                mapped['email_col'] = col
    return mapped

def normalize_fuzzy_program_names(name):
    if not isinstance(name, str):
        return name
    name_norm = name.strip()
    if "Abcess vú" in name_norm:
        return "Chẩn đoán, điều trị và chăm sóc NB áp xe vú"
    return name_norm

def get_status_rank(status):
    if pd.isna(status):
        return 3
    s = str(status).lower()
    if 'không thực hiện' in s:
        return 2
    if 'chưa thực hiện' in s:
        return 1
    return 0

def run_scraper(creds_json, reg_folder_id, main_excel_id):
    # Xác thực Google Drive API
    creds_dict = json.loads(creds_json)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=["https://www.googleapis.com/auth/drive"]
    )
    drive_service = build('drive', 'v3', credentials=creds)

    print("--- BƯỚC 1: ĐỌC FILE EXCEL CHÍNH TỪ GOOGLE DRIVE (KÈM CONVERT ĐỂ LẤY SMART CHIP) ---")
    try:
        # 1. Tạo bản sao tạm thời dưới định dạng Google Sheets để ép Google Drive render Smart Chips
        file_metadata = {
            'name': 'Temp_Dashboard_Sheet_For_Scraping',
            'mimeType': 'application/vnd.google-apps.spreadsheet'
        }
        temp_sheet = drive_service.files().copy(fileId=main_excel_id, body=file_metadata).execute()
        temp_id = temp_sheet['id']
        
        # 2. Export file Google Sheets tạm thời này ngược lại thành .xlsx
        # Quá trình này sẽ ép Google Sheets chuyển đổi Smart Chips thành Hyperlink tiêu chuẩn của Excel
        request = drive_service.files().export_media(
            fileId=temp_id, 
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        main_file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(main_file_stream, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            
        # 3. Xóa file tạm
        try:
            drive_service.files().delete(fileId=temp_id).execute()
        except:
            pass
            
        main_file_stream.seek(0)
    except Exception as e:
        print(f"Lỗi khi convert qua Google Sheets, fallback về tải gốc: {e}")
        request = drive_service.files().get_media(fileId=main_excel_id)
        main_file_stream = io.BytesIO()
        downloader = MediaIoBaseDownload(main_file_stream, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        main_file_stream.seek(0)
    
    # Read main df_sessions
    df_sessions = pd.read_excel(main_file_stream, sheet_name="SHCM - Độ")
    
    print("\n--- BƯỚC 2: QUÉT CÁC FILE TRONG THƯ MỤC ĐĂNG KÝ ---")
    results = drive_service.files().list(
        q=f"'{reg_folder_id}' in parents and trashed=false",
        pageSize=100, fields="nextPageToken, files(id, name)"
    ).execute()
    items = results.get('files', [])
    
    prog_date_map = {}
    student_records = []
    link_cache = {}
    
    # Hàm phụ để tải file từ drive folder
    def download_drive_file(file_id):
        req = drive_service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        dl = MediaIoBaseDownload(fh, req)
        d = False
        while not d:
            _, d = dl.next_chunk()
        fh.seek(0)
        return fh

    for item in items:
        if not item['name'].endswith('.xlsx'): continue
        print(f"Đang xử lý file: {item['name']}")
        
        try:
            fh = download_drive_file(item['id'])
            xl = pd.ExcelFile(fh)
            valid_sheets = [s for s in xl.sheet_names if s.strip().lower() in ('kế hoạch', 'khoa sản', 'khoa phụ')]
            for sheet in valid_sheets:
                fh.seek(0)
                df_temp = pd.read_excel(fh, sheet_name=sheet, nrows=10)
                header_idx = 0
                for idx, row in df_temp.iterrows():
                    row_str = [str(val).lower() for val in row.values]
                    if any('post test' in val or 'gắn link' in val for val in row_str):
                        header_idx = idx + 1
                        break
                
                fh.seek(0)
                df_plan = pd.read_excel(fh, sheet_name=sheet, skiprows=header_idx)
                
                # Tìm cột ngày tháng
                prog_col, day_col, month_col, year_col, unit_col, link_col = None, None, None, None, None, None
                for col in df_plan.columns:
                    col_norm = normalize_text_std(col)
                    if ('nội dung' in col_norm or 'chương trình' in col_norm or 'tên bài giảng' in col_norm) and 'phụ trách' not in col_norm and 'người' not in col_norm:
                        prog_col = col
                    elif col_norm == 'ngày': day_col = col
                    elif col_norm == 'tháng': month_col = col
                    elif col_norm == 'năm': year_col = col
                    elif 'đơn vị' in col_norm or 'khoa' in col_norm: unit_col = col
                    elif ('gắn link post test' in col_norm or 'link post test' in col_norm or 'post test' in col_norm) and 'có' not in col_norm and 'đúng form' not in col_norm:
                        link_col = col
                        
                if link_col is None:
                    for col in df_plan.columns:
                        if 'gắn link' in normalize_text_std(col):
                            link_col = col
                            break
                            
                # Khởi tạo openpyxl để lấy Link chìm
                ws_op = None
                link_col_idx_op = -1
                try:
                    fh.seek(0)
                    wb_op = openpyxl.load_workbook(fh, data_only=False)
                    ws_op = wb_op[sheet]
                    if header_idx + 1 <= ws_op.max_row:
                        for i, cell_op in enumerate(ws_op[header_idx + 1]):
                            if cell_op.value and isinstance(cell_op.value, str):
                                c_norm = normalize_text_std(cell_op.value)
                                if ('gắn link post test' in c_norm or 'link post test' in c_norm or 'post test' in c_norm) and 'có' not in c_norm and 'đúng form' not in c_norm:
                                    link_col_idx_op = i + 1
                                    break
                        if link_col_idx_op == -1:
                            for i, cell_op in enumerate(ws_op[header_idx + 1]):
                                if cell_op.value and isinstance(cell_op.value, str) and 'gắn link' in normalize_text_std(cell_op.value):
                                    link_col_idx_op = i + 1
                                    break
                except:
                    pass
                
                # Quét từng dòng
                for idx, row in df_plan.iterrows():
                    prog_name = str(row.get(prog_col, '')).strip() if prog_col else ''
                    if not prog_name or prog_name.lower() in ('nan', ''): continue
                    prog_clean = clean_program_name(prog_name)
                    
                    # Lấy ngày tháng
                    day = row.get(day_col)
                    month = row.get(month_col)
                    year = row.get(year_col)
                    if pd.notna(day) and pd.notna(month):
                        try:
                            prog_date_map[prog_clean] = (int(float(str(day))), int(float(str(month))), int(float(str(year))) if pd.notna(year) else 2026)
                        except: pass
                        
                    # Lấy link và bóc dữ liệu học viên
                    link_val = row.get(link_col) if link_col else (row.iloc[21] if len(row) > 21 else None)
                    urls = get_urls_from_cell(link_val, drive_service)
                    
                    if not urls and ws_op is not None and link_col_idx_op != -1:
                        excel_row = idx + header_idx + 2
                        try:
                            cell_op = ws_op.cell(row=excel_row, column=link_col_idx_op)
                            if cell_op.hyperlink and cell_op.hyperlink.target:
                                urls = get_urls_from_cell(cell_op.hyperlink.target, drive_service)
                        except: pass
                        
                    if not urls: continue
                    
                    unit = str(row.get(unit_col, '')).strip() if unit_col else ''
                    if not unit or unit.lower() == 'nan':
                        match_prefix = re.match(r'^\[([^\]]+)\]', item['name'])
                        unit = match_prefix.group(1) if match_prefix else item['name'].replace('_Đăng ký SHCM 2026.xlsx', '').replace('.xlsx', '').strip()
                        
                    url = urls[-1]
                    if url in link_cache:
                        df_dl = link_cache[url].copy()
                    else:
                        dl_url, url_type = get_download_url(url)
                        if dl_url:
                            try:
                                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
                                resp = requests.get(dl_url, headers=headers, timeout=20)
                                if resp.status_code == 200:
                                    df_dl = pd.read_excel(io.BytesIO(resp.content))
                                    link_cache[url] = df_dl
                                else:
                                    continue
                            except: continue
                        else:
                            continue
                            
                    df_dl = find_header_and_data(df_dl)
                    col_map = map_columns(df_dl)
                    
                    for i_dl, r_dl in df_dl.iterrows():
                        hv_name = r_dl[col_map['name_col']] if col_map['name_col'] else np.nan
                        hv_id = r_dl[col_map['id_col']] if col_map['id_col'] else np.nan
                        score = r_dl[col_map['score_col']] if col_map['score_col'] else np.nan
                        email = r_dl[col_map['email_col']] if col_map['email_col'] else np.nan
                        
                        if pd.isna(hv_name) and pd.isna(hv_id) and pd.isna(email):
                            continue
                            
                        student_records.append({
                            'Gốc': str(hv_name) if pd.notna(hv_name) else None,
                            'Họ và tên': str(hv_name) if pd.notna(hv_name) else None,
                            'Mã nhân sự': str(hv_id) if pd.notna(hv_id) else None,
                            'Điểm số': str(score) if pd.notna(score) else None,
                            'Địa chỉ email': str(email) if pd.notna(email) else None,
                            'Đơn vị đào tạo': unit,
                            'Ngày': day,
                            'Tháng': month,
                            'Năm': year if pd.notna(year) else 2026,
                            'Tên chương trình': prog_name
                        })
                        
        except Exception as e:
            print(f"Lỗi: {e}")

    print("\n--- BƯỚC 3: CẬP NHẬT FILE CHÍNH ---")
    main_file_stream.seek(0)
    wb = openpyxl.load_workbook(main_file_stream, data_only=False)
    
    import datetime
    print("--- BƯỚC 2.5: BÓC TÁCH LINK TỪ FILE CHÍNH ---")
    ws_main = wb["SHCM - Độ"] if "SHCM - Độ" in wb.sheetnames else wb.worksheets[0]
    
    try:
        sheets_service = build('sheets', 'v4', credentials=creds)
        sheet_data = sheets_service.spreadsheets().get(
            spreadsheetId=main_excel_id, 
            includeGridData=True, 
            ranges=["SHCM - Độ"]
        ).execute()
        grid_data = sheet_data.get('sheets', [{}])[0].get('data', [{}])[0].get('rowData', [])
    except Exception as e:
        print(f"Lỗi Sheets API: {e}")
        grid_data = []
    
    prog_col_idx = -1
    link_col_idx = -1
    unit_col_idx = -1
    day_col_idx, month_col_idx, year_col_idx = -1, -1, -1
    date_col_idx = -1
    
    for i, cell in enumerate(ws_main[1]):
        if not cell.value: continue
        val = normalize_text_std(cell.value)
        if 'nội dung' in val or 'chương trình' in val: prog_col_idx = i + 1
        elif 'gắn link' in val and 'post test' in val: link_col_idx = i + 1
        elif 'đơn vị' in val: unit_col_idx = i + 1
        elif val == 'ngày': day_col_idx = i + 1
        elif val == 'tháng': month_col_idx = i + 1
        elif val == 'năm': year_col_idx = i + 1
        elif 'ngày cụ thể' in val: date_col_idx = i + 1
            
    if prog_col_idx != -1 and link_col_idx != -1:
        for row_idx in range(2, ws_main.max_row + 1):
            prog_name = ws_main.cell(row=row_idx, column=prog_col_idx).value
            if not prog_name: continue
            
            link_cell = ws_main.cell(row=row_idx, column=link_col_idx)
            urls = []
            
            api_url = None
            if grid_data and (row_idx - 1) < len(grid_data):
                row_data = grid_data[row_idx - 1].get('values', [])
                if (link_col_idx - 1) < len(row_data):
                    cell_data = row_data[link_col_idx - 1]
                    if 'hyperlink' in cell_data:
                        api_url = cell_data['hyperlink']
                    elif 'formattedValue' in cell_data and 'http' in str(cell_data['formattedValue']):
                        api_url = cell_data['formattedValue']
                        
            if api_url:
                urls = get_urls_from_cell(api_url, drive_service)
                
            if not urls:
                if link_cell.hyperlink and link_cell.hyperlink.target:
                    urls = get_urls_from_cell(link_cell.hyperlink.target, drive_service)
                elif link_cell.value:
                    urls = get_urls_from_cell(link_cell.value, drive_service)
                
            if not urls: continue
            url = urls[-1]
            if url in link_cache: df_dl = link_cache[url].copy()
            else:
                dl_url, url_type = get_download_url(url)
                if dl_url:
                    try:
                        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
                        resp = requests.get(dl_url, headers=headers, timeout=20)
                        if resp.status_code == 200:
                            df_dl = pd.read_excel(io.BytesIO(resp.content))
                            link_cache[url] = df_dl
                        else: continue
                    except: continue
                else: continue
                
            df_dl = find_header_and_data(df_dl)
            col_map = map_columns(df_dl)
            
            unit = ws_main.cell(row=row_idx, column=unit_col_idx).value if unit_col_idx != -1 else ''
            day, month, year = np.nan, np.nan, 2026
            
            if date_col_idx != -1:
                date_val = ws_main.cell(row=row_idx, column=date_col_idx).value
                if isinstance(date_val, datetime.datetime):
                    day, month, year = date_val.day, date_val.month, date_val.year
            else:
                try:
                    day = ws_main.cell(row=row_idx, column=day_col_idx).value if day_col_idx != -1 else np.nan
                    month = ws_main.cell(row=row_idx, column=month_col_idx).value if month_col_idx != -1 else np.nan
                    year = ws_main.cell(row=row_idx, column=year_col_idx).value if year_col_idx != -1 else 2026
                except: pass
                
            for i_dl, r_dl in df_dl.iterrows():
                hv_name = r_dl[col_map['name_col']] if col_map['name_col'] else np.nan
                hv_id = r_dl[col_map['id_col']] if col_map['id_col'] else np.nan
                score = r_dl[col_map['score_col']] if col_map['score_col'] else np.nan
                email = r_dl[col_map['email_col']] if col_map['email_col'] else np.nan
                
                if pd.isna(hv_name) and pd.isna(hv_id) and pd.isna(email): continue
                    
                student_records.append({
                    'Gốc': str(hv_name) if pd.notna(hv_name) else None,
                    'Họ và tên': str(hv_name) if pd.notna(hv_name) else None,
                    'Mã nhân sự': str(hv_id) if pd.notna(hv_id) else None,
                    'Điểm số': str(score) if pd.notna(score) else None,
                    'Địa chỉ email': str(email) if pd.notna(email) else None,
                    'Đơn vị đào tạo': str(unit) if unit else '',
                    'Ngày': day,
                    'Tháng': month,
                    'Năm': year,
                    'Tên chương trình': str(prog_name)
                })

    df_students = pd.DataFrame(student_records)
    
    if 'Dữ liệu học viên' in wb.sheetnames:
        del wb['Dữ liệu học viên']
    ws_stud = wb.create_sheet('Dữ liệu học viên')
    
    if not df_students.empty:
        ws_stud.append(df_students.columns.tolist())
        df_students_clean = df_students.replace({np.nan: None})
        for r in df_students_clean.itertuples(index=False):
            ws_stud.append(list(r))
            
    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    
    media = MediaIoBaseUpload(out_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
    drive_service.files().update(fileId=main_excel_id, media_body=media).execute()
    print("Cập nhật thành công lên Google Drive!")
    
    return True
