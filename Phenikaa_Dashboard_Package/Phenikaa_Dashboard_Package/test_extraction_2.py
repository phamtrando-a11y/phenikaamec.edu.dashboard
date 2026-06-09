import openpyxl

wb = openpyxl.load_workbook('../250903_Theo dõi cấp CME + Điểm danh_Bài giảng SHCM (Độ)_unlocked.xlsx', data_only=False)
ws = wb['SHCM - Độ']

prog_col_idx = 8
link_col_idx = 20

for row_idx in range(2, 50):
    prog_name = ws.cell(row=row_idx, column=prog_col_idx).value
    if prog_name and "cấp cứu" in str(prog_name).lower():
        link_cell = ws.cell(row=row_idx, column=link_col_idx)
        print(f"Prog: {prog_name}")
        print(f"Value: {repr(link_cell.value)}")
        print(f"Hyperlink target: {link_cell.hyperlink.target if link_cell.hyperlink else None}")
