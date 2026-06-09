import pandas as pd
import openpyxl
from google_drive_scraper import get_urls_from_cell, get_download_url

wb = openpyxl.load_workbook('../250903_Theo dõi cấp CME + Điểm danh_Bài giảng SHCM (Độ)_unlocked.xlsx', data_only=False)
ws = wb['SHCM - Độ']

prog_col_idx = 8
link_col_idx = 20

for row_idx in range(2, 50):
    prog_name = ws.cell(row=row_idx, column=prog_col_idx).value
    if prog_name and "cấp cứu" in str(prog_name).lower():
        link_cell = ws.cell(row=row_idx, column=link_col_idx)
        print(f"Prog: {prog_name}")
        print(f"Value: {link_cell.value}")
        print(f"Hyperlink target: {link_cell.hyperlink.target if link_cell.hyperlink else None}")
        
        urls = []
        if link_cell.hyperlink and link_cell.hyperlink.target:
            urls = get_urls_from_cell(link_cell.hyperlink.target, None)
        elif link_cell.value:
            urls = get_urls_from_cell(link_cell.value, None)
            
        print("URLs extracted:", urls)
        
        if urls:
            dl_url, _ = get_download_url(urls[-1])
            print("Download URL:", dl_url)
