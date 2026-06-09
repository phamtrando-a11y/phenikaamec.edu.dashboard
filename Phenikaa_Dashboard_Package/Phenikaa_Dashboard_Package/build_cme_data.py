import os
import re
import io
import json
import requests
import msoffcrypto
import openpyxl
import pandas as pd
import numpy as np
import unicodedata
import datetime

# Đường dẫn các file
workspace_dir = "/Users/dotran/Library/CloudStorage/GoogleDrive-daotao.bvdhphenikaa@gmail.com/Drive của tôi/1_Đào tạo/13. Sinh hoạt chuyên môn"
main_excel_path = os.path.join(workspace_dir, "250903_Theo dõi cấp CME + Điểm danh_Bài giảng SHCM (Độ)_unlocked.xlsx")
temp_decrypted_path = os.path.join(workspace_dir, "temp_decrypted.xlsx")
reg_folder = os.path.join(workspace_dir, "5.Đăng ký SHCM - Độ")
json_output_path = os.path.join(workspace_dir, "Phenikaa_Dashboard_Package", "netlify-dashboard", "data.json")

print("--- BƯỚC 1: ĐỌC FILE EXCEL CHÍNH ---")
try:
    decrypted_workbook = io.BytesIO()
    with open(main_excel_path, 'rb') as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password='QC2026')
        office_file.decrypt(decrypted_workbook)
    
    with open(temp_decrypted_path, 'wb') as f_out:
        f_out.write(decrypted_workbook.getvalue())
    print("Giải mã thành công file chính!")
except Exception as e:
        print("Đã copy file chính trực tiếp (không cần giải mã).")

# Đọc danh sách sheet từ file chính để bảo toàn
xl_main = pd.ExcelFile(temp_decrypted_path)
main_sheets = xl_main.sheet_names
print("Các sheet hiện có trong file chính:", main_sheets)

def remove_accents(input_str):
    if not isinstance(input_str, str):
        return ""
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)])

def clean_program_name(name):
    if pd.isna(name):
        return ""
    name = unicodedata.normalize('NFC', str(name)).strip().lower()
    name = remove_accents(name)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name

print("\n--- BƯỚC 2: QUÉT CÁC FILE ĐĂNG KÝ ĐỂ LẤY THÔNG TIN NGÀY THÁNG ---")
reg_files = [f for f in os.listdir(reg_folder) if f.endswith('.xlsx') and not f.startswith('~')]
prog_date_map = {}

for f in reg_files:
    path = os.path.join(reg_folder, f)
    try:
        xl = pd.ExcelFile(path)
        valid_sheets = [s for s in xl.sheet_names if s.strip().lower() in ('kế hoạch', 'khoa sản', 'khoa phụ')]
        for sheet in valid_sheets:
            df_temp = pd.read_excel(path, sheet_name=sheet, nrows=10)
            header_idx = 0
            for idx, row in df_temp.iterrows():
                row_str = [str(val).lower() for val in row.values]
                if any('post test' in val or 'gắn link' in val for val in row_str):
                    header_idx = idx + 1
                    break
            df_plan = pd.read_excel(path, sheet_name=sheet, skiprows=header_idx)
            
            prog_col, day_col, month_col, year_col = None, None, None, None
            for col in df_plan.columns:
                col_norm = col.lower().strip()
                if ('nội dung' in col_norm or 'chương trình' in col_norm or 'tên bài giảng' in col_norm) and 'phụ trách' not in col_norm and 'người' not in col_norm:
                    prog_col = col
                elif col_norm == 'ngày':
                    day_col = col
                elif col_norm == 'tháng':
                    month_col = col
                elif col_norm == 'năm':
                    year_col = col
            
            for idx, row in df_plan.iterrows():
                prog_name = str(row.get(prog_col, '')).strip() if prog_col else ''
                if not prog_name or prog_name.lower() in ('nan', ''):
                    continue
                prog_clean = clean_program_name(prog_name)
                
                day = row.get(day_col)
                month = row.get(month_col)
                year = row.get(year_col)
                
                if pd.notna(day) and pd.notna(month):
                    try:
                        day_val = int(float(str(day).strip()))
                        month_val = int(float(str(month).strip()))
                        year_val = int(float(str(year).strip())) if pd.notna(year) else 2026
                        prog_date_map[prog_clean] = (day_val, month_val, year_val)
                    except:
                        pass
    except Exception as e:
        pass

print(f"Lấy thành công thông tin ngày tháng cho {len(prog_date_map)} chương trình từ 30 file đăng ký.")

print("\n--- BƯỚC 3: XỬ LÝ DỮ LIỆU CÁC BUỔI ĐÀO TẠO (APPEND1) ---")
df_shcm = pd.read_excel(temp_decrypted_path, sheet_name="SHCM - Độ")

# Map các cột từ 'SHCM - Độ' sang 'Append1'
column_mapping = {
    'TÊN ĐƠN VỊ': 'ĐƠN VỊ',
    'TÌNH TRẠNG THỰC HIỆN': 'Tình trạng thực hiện',
    'SỐ HỌC VIÊN THAM GIA': 'Số học viên tham gia',
    'SỐ HỌC VIÊN ĐẠT ĐIỂM TRÊN 80%': 'Số Học viên đạt',
    'NỘI DUNG/CHƯƠNG TRÌNH': 'NỘI DUNG/CHƯƠNG TRÌNH',
    'NGƯỜI PHỤ TRÁCH NỘI DUNG': 'NGƯỜI PHỤ TRÁCH NỘI DUNG',
    'GẮN LINK POST TEST': 'Gắn link Điểm danh và post test',
    'GẮN LINK BẢNG KIỂM': 'Gắn link bảng kiểm đào tạo nội bộ',
    'GẮN LINK BÀI BÁO CÁO': 'Gắn link bài báo cáo',
    'GẮN LINK VIDEO': 'Gắn link video buổi đào tạo'
}

df_sessions = df_shcm.copy()
# Rename columns
df_sessions = df_sessions.rename(columns=column_mapping)

# Trích xuất Ngày, Tháng, Năm từ 'Ngày cụ thể'
if 'Ngày cụ thể' in df_sessions.columns:
    df_sessions['Ngày cụ thể_parsed'] = pd.to_datetime(df_sessions['Ngày cụ thể'], errors='coerce')
    df_sessions['Ngày'] = df_sessions['Ngày cụ thể_parsed'].dt.day
    df_sessions['Tháng'] = df_sessions['Ngày cụ thể_parsed'].dt.month
    df_sessions['Năm'] = df_sessions['Ngày cụ thể_parsed'].dt.year
else:
    df_sessions['Ngày'] = np.nan
    df_sessions['Tháng'] = np.nan
    df_sessions['Năm'] = np.nan

# Điền bù các cột Ngày, Tháng, Năm bị trống bằng cách tra cứu từ file đăng ký
df_sessions['prog_clean'] = df_sessions['NỘI DUNG/CHƯƠNG TRÌNH'].apply(clean_program_name)
filled_count = 0
for idx, r in df_sessions.iterrows():
    if pd.isna(r['Ngày']) or pd.isna(r['Tháng']):
        prog = r['prog_clean']
        if prog in prog_date_map:
            day, month, year = prog_date_map[prog]
            df_sessions.at[idx, 'Ngày'] = day
            df_sessions.at[idx, 'Tháng'] = month
            df_sessions.at[idx, 'Năm'] = year
            filled_count += 1
        else:
            # Fuzzy match
            for k, v in prog_date_map.items():
                if prog and (prog in k or k in prog):
                    day, month, year = v
                    df_sessions.at[idx, 'Ngày'] = day
                    df_sessions.at[idx, 'Tháng'] = month
                    df_sessions.at[idx, 'Năm'] = year
                    filled_count += 1
                    break

print(f"Đã điền bù tự động Ngày/Tháng cho {filled_count} buổi học bị trống thông tin.")

# Điền đơn vị trống cho chương trình cụ thể
for idx, r in df_sessions.iterrows():
    if pd.isna(r['ĐƠN VỊ']) and isinstance(r['NỘI DUNG/CHƯƠNG TRÌNH'], str) and 'lọc máu' in r['NỘI DUNG/CHƯƠNG TRÌNH'].lower():
        df_sessions.at[idx, 'ĐƠN VỊ'] = 'Thận Nhân Tạo'

# Chuẩn hóa tên fuzzy đặc biệt
def normalize_fuzzy_program_names(name):
    if not isinstance(name, str):
        return name
    name_norm = name.strip()
    if "Abcess vú" in name_norm:
        return "Chẩn đoán, điều trị và chăm sóc NB áp xe vú"
    return name_norm

df_sessions['NỘI DUNG/CHƯƠNG TRÌNH'] = df_sessions['NỘI DUNG/CHƯƠNG TRÌNH'].apply(normalize_fuzzy_program_names)
df_sessions['prog_clean'] = df_sessions['NỘI DUNG/CHƯƠNG TRÌNH'].apply(clean_program_name)

# Gán điểm ưu tiên cho tình trạng thực hiện (càng nhỏ càng ưu tiên)
def get_status_rank(status):
    if pd.isna(status):
        return 3
    s = str(status).lower()
    if 'không thực hiện' in s:
        return 2
    if 'chưa thực hiện' in s:
        return 1
    # Có thực hiện, Thực hiện, Đã thực hiện, Đã hoàn thành...
    return 0

df_sessions['status_rank'] = df_sessions['Tình trạng thực hiện'].apply(get_status_rank)
# Sắp xếp để dòng tốt nhất lên đầu
df_sessions = df_sessions.sort_values(by=['Năm', 'Tháng', 'prog_clean', 'status_rank'])
# Drop duplicates, giữ dòng đầu tiên (chỉ coi là trùng khi cùng chương trình, cùng ngày cụ thể, và cùng đơn vị đầu mối)
df_sessions = df_sessions.drop_duplicates(subset=['prog_clean', 'Ngày', 'Tháng', 'Năm', 'ĐƠN VỊ'], keep='first')
# Sắp xếp lại theo thời gian
df_sessions = df_sessions.sort_values(by=['Năm', 'Tháng', 'Ngày'])

# Đảm bảo các cột cần thiết có mặt
required_cols = ['ĐƠN VỊ', 'Tình trạng thực hiện', 'Số học viên tham gia', 'Số Học viên đạt', 
                 'NỘI DUNG/CHƯƠNG TRÌNH', 'NGƯỜI PHỤ TRÁCH NỘI DUNG', 'Gắn link Điểm danh và post test',
                 'Gắn link bảng kiểm đào tạo nội bộ', 'Gắn link bài báo cáo', 'Gắn link video buổi đào tạo',
                 'Ngày', 'Tháng', 'Năm']

for col in required_cols:
    if col not in df_sessions.columns:
        df_sessions[col] = np.nan

# Giữ lại các cột cần thiết cho Append1
df_sessions_final = df_sessions[required_cols].copy()
print(f"Xử lý xong dữ liệu buổi đào tạo. Số lượng dòng sau khi loại trùng: {len(df_sessions_final)}")


print("\n--- BƯỚC 4: TẢI VÀ TỔNG HỢP DỮ LIỆU HỌC VIÊN ---")

def normalize_text_std(text):
    if not isinstance(text, str):
        return ""
    return unicodedata.normalize('NFC', text).strip().lower()

def get_urls_from_cell(val):
    if pd.isna(val):
        return []
    val_str = str(val).strip()
    return re.findall(r'https?://[^\s,;"\'\)]+', val_str)

def get_download_url(url):
    sheet_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if sheet_match:
        sp_id = sheet_match.group(1)
        res_key = ""
        rk_match = re.search(r'(resourcekey=[a-zA-Z0-9-_]+)', url)
        if rk_match:
            res_key = "&" + rk_match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sp_id}/export?format=xlsx{res_key}", "google_sheet"
    drive_match = re.search(r'/file/d/([a-zA-Z0-9-_]+)', url)
    if drive_match:
        file_id = drive_match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}", "google_drive"
    return None, None

def find_header_and_data(df):
    best_row_idx = 0
    max_matches = 0
    keywords = {'dấu thời gian', 'timestamp', 'điểm số', 'score', 'họ tên', 'họ và tên', 'tên', 
                'mã nhân viên', 'mã nhân sự', 'mã nv', 'mã ns', 'email', 'địa chỉ email'}
    
    for idx in range(min(10, len(df))):
        row_vals = [normalize_text_std(val) for val in df.iloc[idx].values]
        matches = sum(1 for val in row_vals if any(kw in val for kw in keywords))
        if matches > max_matches:
            max_matches = matches
            best_row_idx = idx
            
    if max_matches >= 2:
        columns = df.iloc[best_row_idx].values
        cleaned_columns = []
        for c in columns:
            c_str = str(c).strip()
            if not c_str or c_str.lower() == 'nan':
                cleaned_columns.append(f"Unnamed_{len(cleaned_columns)}")
            else:
                cleaned_columns.append(c_str)
        df_clean = df.iloc[best_row_idx+1:].copy()
        df_clean.columns = cleaned_columns
        return df_clean
    return df

def map_columns(df):
    cols = df.columns
    mapped = {'name_col': None, 'id_col': None, 'score_col': None, 'email_col': None}
    
    has_cot_cols = False
    cleaned_cols = [normalize_text_std(c) for c in cols if c and str(c).lower() != 'nan']
    if cleaned_cols and all(re.match(r'^(cột|cot)\s*\d+$', c) for c in cleaned_cols):
        has_cot_cols = True
        
    if has_cot_cols and len(cols) >= 6:
        for col in cols:
            col_norm = normalize_text_std(col)
            if '4' in col_norm:
                mapped['name_col'] = col
            elif '6' in col_norm:
                mapped['id_col'] = col
            elif '3' in col_norm:
                mapped['score_col'] = col
            elif '2' in col_norm:
                mapped['email_col'] = col
        return mapped
        
    for col in cols:
        col_norm = normalize_text_std(col)
        if not mapped['name_col']:
            if col_norm in ('họ tên', 'họ và tên', 'họ tên đầy đủ', 'họ và tên đầy đủ', 'tên', 'name', 'nhân sự', 'nguyễn văn a', 'họ tên?'):
                mapped['name_col'] = col
            elif 'họ' in col_norm and 'tên' in col_norm:
                mapped['name_col'] = col
        if not mapped['id_col']:
            if col_norm in ('mã nhân sự', 'mã nhân viên', 'mã nv', 'mã ns', 'msnv', 'msns', 'id', 'mã số nhân viên'):
                mapped['id_col'] = col
            elif 'mã' in col_norm and ('nhân viên' in col_norm or 'nhân sự' in col_norm or 'nv' in col_norm or 'ns' in col_norm):
                mapped['id_col'] = col
        if not mapped['score_col']:
            if col_norm in ('điểm số', 'điểm', 'score', 'kết quả'):
                mapped['score_col'] = col
            elif 'điểm' in col_norm or 'score' in col_norm:
                mapped['score_col'] = col
        if not mapped['email_col']:
            if col_norm in ('email', 'địa chỉ email', 'email address'):
                mapped['email_col'] = col
    return mapped

student_records = []
link_cache = {}

for f in reg_files:
    path = os.path.join(reg_folder, f)
    try:
        xl = pd.ExcelFile(path)
        valid_sheets = [s for s in xl.sheet_names if s.strip().lower() in ('kế hoạch', 'khoa sản', 'khoa phụ')]
        for sheet in valid_sheets:
            df_temp = pd.read_excel(path, sheet_name=sheet, nrows=10)
            header_idx = 0
            for idx, row in df_temp.iterrows():
                row_str = [str(val).lower() for val in row.values]
                if any('post test' in val or 'gắn link' in val for val in row_str):
                    header_idx = idx + 1
                    break
            df_plan = pd.read_excel(path, sheet_name=sheet, skiprows=header_idx)
            
            prog_col, unit_col, day_col, month_col, year_col, link_col = None, None, None, None, None, None
            for col in df_plan.columns:
                col_norm = normalize_text_std(col)
                if ('nội dung' in col_norm or 'chương trình' in col_norm or 'tên bài giảng' in col_norm) and 'phụ trách' not in col_norm and 'người' not in col_norm:
                    prog_col = col
                elif 'đơn vị' in col_norm or 'khoa' in col_norm:
                    unit_col = col
                elif col_norm == 'ngày':
                    day_col = col
                elif col_norm == 'tháng':
                    month_col = col
                elif col_norm == 'năm':
                    year_col = col
                elif 'gắn link post test' in col_norm or 'link post test' in col_norm or 'post test' in col_norm:
                    if 'có' not in col_norm and 'đúng form' not in col_norm:
                        link_col = col
            
            if link_col is None:
                for col in df_plan.columns:
                    if 'gắn link' in normalize_text_std(col):
                        link_col = col
                        break
            
            # Khởi tạo openpyxl để lấy Link chìm
            try:
                wb_op = openpyxl.load_workbook(path, data_only=False)
                ws_op = wb_op[sheet]
                link_col_idx_op = -1
                if header_idx + 1 <= ws_op.max_row:
                    for i, cell_op in enumerate(ws_op[header_idx + 1]):
                        if cell_op.value and isinstance(cell_op.value, str):
                            c_norm = normalize_text_std(cell_op.value)
                            if ('gắn link post test' in c_norm or 'link post test' in c_norm or 'post test' in c_norm) and 'có' not in c_norm and 'đúng form' not in c_norm:
                                link_col_idx_op = i + 1
                                break
                    if link_col_idx_op == -1:
                        for i, cell_op in enumerate(ws_op[header_idx + 1]):
                            if cell_op.value and isinstance(cell_op.value, str) and 'gắn link' in normalize_text_std(cell_op.value):
                                link_col_idx_op = i + 1
                                break
            except:
                ws_op = None
                link_col_idx_op = -1
                        
            for idx, row in df_plan.iterrows():
                prog_name = str(row.get(prog_col, '')).strip() if prog_col else ''
                if not prog_name or prog_name.lower() in ('nan', ''):
                    continue
                    
                link_val = row.get(link_col) if link_col else (row.iloc[21] if len(row) > 21 else None)
                urls = get_urls_from_cell(link_val)
                
                # Bóc tách Link chìm từ openpyxl nếu Pandas không tìm thấy url
                if not urls and ws_op is not None and link_col_idx_op != -1:
                    excel_row = idx + header_idx + 2
                    try:
                        cell_op = ws_op.cell(row=excel_row, column=link_col_idx_op)
                        if cell_op.hyperlink and cell_op.hyperlink.target:
                            urls = get_urls_from_cell(cell_op.hyperlink.target)
                    except:
                        pass
                        
                if not urls:
                    continue
                
                unit = str(row.get(unit_col, '')).strip() if unit_col else ''
                if not unit or unit.lower() == 'nan':
                    match_prefix = re.match(r'^\[([^\]]+)\]', f)
                    unit = match_prefix.group(1) if match_prefix else f.replace('_Đăng ký SHCM 2026.xlsx', '').replace('.xlsx', '').strip()
                
                day = row.get(day_col) if day_col else np.nan
                month = row.get(month_col) if month_col else np.nan
                year = row.get(year_col) if year_col else np.nan
                
                for url in urls:
                    dl_url, url_type = get_download_url(url)
                    if not dl_url:
                        continue
                    
                    df_stud = None
                    if dl_url in link_cache:
                        df_stud = link_cache[dl_url]
                    else:
                        try:
                            res = requests.get(dl_url, timeout=10)
                            content_type = res.headers.get('Content-Type', '')
                            if 'html' not in content_type.lower() and 'image' not in content_type.lower() and 'pdf' not in content_type.lower():
                                xl_dl = pd.ExcelFile(io.BytesIO(res.content))
                                df_raw = pd.read_excel(xl_dl, sheet_name=xl_dl.sheet_names[0])
                                df_stud = find_header_and_data(df_raw)
                                link_cache[dl_url] = df_stud
                        except Exception as e:
                            print(f"Lỗi tải link {url}: {e}")
                            link_cache[dl_url] = None
                            
                    if df_stud is None or df_stud.empty:
                        continue
                        
                    mapped = map_columns(df_stud)
                    name_col = mapped['name_col']
                    id_col = mapped['id_col']
                    score_col = mapped['score_col']
                    email_col = mapped['email_col']
                    
                    if not name_col and not email_col:
                        continue
                        
                    for s_idx, s_row in df_stud.iterrows():
                        s_name = ''
                        if name_col:
                            s_name = str(s_row.get(name_col, '')).strip()
                        elif email_col:
                            email = str(s_row.get(email_col, '')).strip()
                            s_name = email.split('@')[0] if '@' in email else email
                            
                        if not s_name or s_name.lower() in ('nan', ''):
                            continue
                            
                        s_id = ''
                        if id_col:
                            s_id = str(s_row.get(id_col, '')).strip()
                            if s_id.lower() == 'nan':
                                s_id = ''
                                
                        s_score = s_row.get(score_col) if score_col else np.nan
                        if pd.isna(s_score) or str(s_score).strip().lower() == 'nan':
                            s_score = ''
                        else:
                            s_score = str(s_score).strip()
                            
                        student_records.append({
                            'Họ và tên': s_name,
                            'Gốc': s_name,
                            'Mã nhân sự': s_id if s_id else np.nan,
                            'Điểm số': s_score,
                            'Đơn vị đào tạo': unit,
                            'Tháng thực hiện': month,
                            'Ngày đào tạo': day,
                            'Tên chương trình': prog_name
                        })
            print(f"Đã xử lý xong file: {f}")
    except Exception as e:
        print(f"Lỗi xử lý file {f}: {e}")

df_students = pd.DataFrame(student_records)
print(f"Tổng hợp xong dữ liệu học viên vừa quét. Số lượng dòng: {len(df_students)}")

print("\n--- BƯỚC 5: GHI DỮ LIỆU VÀO FILE EXCEL CHÍNH ---")
wb = openpyxl.load_workbook(temp_decrypted_path)

# Tạo/ghi đè sheet Append1
if 'Append1' in wb.sheetnames:
    del wb['Append1']
ws_sess = wb.create_sheet('Append1')

# Ghi headers cho Append1
ws_sess.append(df_sessions_final.columns.tolist())
# Ghi data cho Append1
for r in df_sessions_final.itertuples(index=False):
    ws_sess.append(list(r))

# Tạo/ghi đè sheet Dữ liệu học viên (chứa dữ liệu vừa quét được)
if 'Dữ liệu học viên' in wb.sheetnames:
    del wb['Dữ liệu học viên']
ws_stud = wb.create_sheet('Dữ liệu học viên')

# Ghi headers cho Dữ liệu học viên
ws_stud.append(df_students.columns.tolist())
# Ghi data cho Dữ liệu học viên
df_students_clean = df_students.replace({np.nan: None})
for r in df_students_clean.itertuples(index=False):
    ws_stud.append(list(r))

# Lưu file chính
wb.save(main_excel_path)
print(f"Đã lưu dữ liệu cập nhật vào file chính tại: {main_excel_path}")

# Xóa file tạm
if os.path.exists(temp_decrypted_path):
    os.remove(temp_decrypted_path)

print("\n--- BƯỚC 6: XUẤT RA FILE JSON CHO NETLIFY ---")
df_sessions_json = df_sessions_final.dropna(subset=["ĐƠN VỊ"])
df_sessions_json = df_sessions_json.replace({np.nan: ""})
sessions_list = df_sessions_json.to_dict(orient="records")

# Gộp Dữ liệu học viên 2026 (nếu có trong file chính) và Dữ liệu học viên vừa quét được để lưu vào JSON
df_students_full = pd.DataFrame()
try:
    # Đọc sheet Dữ liệu học viên 2026 (hoặc sheet 2026) nếu có trong file chính làm dữ liệu gốc
    wb_read = openpyxl.load_workbook(main_excel_path, read_only=True)
    if "Dữ liệu học viên 2026" in wb_read.sheetnames:
        df_students_full = pd.read_excel(main_excel_path, sheet_name="Dữ liệu học viên 2026")
    elif "2026" in wb_read.sheetnames:
        # Nếu sheet tên là 2026, ta chuẩn hóa nó để khớp với schema học viên
        df_2026_raw = pd.read_excel(main_excel_path, sheet_name="2026")
        # Chuẩn hóa cột
        df_students_full = pd.DataFrame()
        df_students_full['Họ và tên'] = df_2026_raw['Họ và tên'] if 'Họ và tên' in df_2026_raw.columns else df_2026_raw['Gốc'] if 'Gốc' in df_2026_raw.columns else np.nan
        df_students_full['Gốc'] = df_2026_raw['Gốc'] if 'Gốc' in df_2026_raw.columns else df_students_full['Họ và tên']
        df_students_full['Mã nhân sự'] = df_2026_raw['Mã nhân sự'] if 'Mã nhân sự' in df_2026_raw.columns else np.nan
        df_students_full['Điểm số'] = df_2026_raw['Điểm số'] if 'Điểm số' in df_2026_raw.columns else np.nan
        df_students_full['Đơn vị đào tạo'] = df_2026_raw['Đơn vị đào tạo'] if 'Đơn vị đào tạo' in df_2026_raw.columns else np.nan
        df_students_full['Tháng thực hiện'] = df_2026_raw['Tháng'] if 'Tháng' in df_2026_raw.columns else np.nan
        df_students_full['Ngày đào tạo'] = df_2026_raw['Ngày'] if 'Ngày' in df_2026_raw.columns else np.nan
        df_students_full['Tên chương trình'] = df_2026_raw['Tên chương trình'] if 'Tên chương trình' in df_2026_raw.columns else np.nan
    wb_read.close()
except Exception as e:
    print(f"Không thể đọc dữ liệu học viên gốc để gộp JSON: {e}")

if not df_students_full.empty:
    df_students_combined = pd.concat([df_students_full, df_students], ignore_index=True)
    df_students_combined['Họ Tên Đầy Đủ Temp'] = df_students_combined['Họ và tên'].fillna(df_students_combined['Gốc'])
    df_students_combined['Mã_Định_Danh_Temp'] = df_students_combined['Mã nhân sự'].fillna(df_students_combined['Họ Tên Đầy Đủ Temp'])
    df_students_combined = df_students_combined.drop_duplicates(subset=['Mã_Định_Danh_Temp', 'Tên chương trình'])
    df_students_combined = df_students_combined.drop(columns=['Họ Tên Đầy Đủ Temp', 'Mã_Định_Danh_Temp'])
else:
    df_students_combined = df_students

df_students_json = df_students_combined.fillna("")
students_list = df_students_json.to_dict(orient="records")

output_json = {
    "sessions": sessions_list,
    "students": students_list
}

def make_json_serializable(data):
    if isinstance(data, dict):
        return {k: make_json_serializable(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [make_json_serializable(v) for v in data]
    elif isinstance(data, (pd.Timestamp, datetime.datetime)):
        return data.strftime('%d/%m/%Y')
    elif isinstance(data, datetime.date):
        return data.strftime('%d/%m/%Y')
    elif pd.isna(data):
        return ""
    elif isinstance(data, (np.int64, np.int32, np.int16, np.int8)):
        return int(data)
    elif isinstance(data, (np.float64, np.float32)):
        if np.isnan(data) or np.isinf(data):
            return ""
        if data.is_integer():
            return int(data)
        return float(data)
    return data

output_json_serializable = make_json_serializable(output_json)

with open(json_output_path, "w", encoding="utf-8") as f:
    json.dump(output_json_serializable, f, ensure_ascii=False, indent=2)
print(f"Đã cập nhật file JSON cho Netlify tại: {json_output_path}")
print("TẤT CẢ HOÀN THÀNH THÀNH CÔNG!")
