import openpyxl

wb = openpyxl.load_workbook('live_excel_real.xlsx', data_only=False)
ws = wb['SHCM - Độ']

prog_col_idx = 8
link_col_idx = 20

for row_idx in range(2, 60):
    prog_name = ws.cell(row=row_idx, column=prog_col_idx).value
    if prog_name and "cấp cứu" in str(prog_name).lower():
        link_cell = ws.cell(row=row_idx, column=link_col_idx)
        print(f"Row {row_idx}: Prog='{prog_name}'")
        print(f"  value: {repr(link_cell.value)}")
        print(f"  hyperlink target: {link_cell.hyperlink.target if link_cell.hyperlink else None}")
